# pyrefly: ignore [missing-import]
"""
SANJIVANI ATTENDANCE — AI-Powered Face Recognition & Smart Attendance Management
Visual Quality Upgrade: Executive Navy & Royal Blue Theme, Refined Typography,
Modern Cards, Table Styling, Interactive Button Hover States, High-DPI Windows Support.
"""

import tkinter as tk
from tkinter import ttk, messagebox
import cv2
import os
import sys
import traceback
import csv
import time
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def handle_exception(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    err_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    try:
        with open("crash_log.txt", "w", encoding="utf-8") as f:
            f.write(err_msg)
    except Exception:
        pass

sys.excepthook = handle_exception

# Set working directory to the application directory (for standalone executable and script execution)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
    try:
        os.chdir(BASE_DIR)
    except Exception:
        pass
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Enable Windows High-DPI awareness for crisp typography and borders
try:
    import ctypes
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    try:
        import ctypes
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

# Set explicit Windows AppUserModelID for consistent Taskbar icon display and grouping
try:
    import ctypes
    app_user_model_id = "Sanjivani.Attendance.System.2.0"
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(app_user_model_id)
except Exception:
    pass

import send_email


# ============================================================
# CONFIGURATION & CONSTANTS (Backend Preserved)
# ============================================================
STUDENT_FOLDER = "students"
ATTENDANCE_CSV = "attendance.csv"
ATTENDANCE_XLSX = "attendance.xlsx"
CASCADE_FILE = "haarcascade_frontalface_default.xml"
FACE_SIZE = (200, 200)

# LBPH confidence threshold for strict Unknown Face Rejection:
#   0 - 55  = Genuine Registered Student (Strong Match)
#   70+     = UNKNOWN FACE REJECTED (Access Denied)
CONFIDENCE_THRESHOLD = 70


# ============================================================
# BACKEND AI & FILE MANAGEMENT (100% Preserved)
# ============================================================

def get_resource_path(relative_path):
    """Get absolute path to bundled static resource (e.g. haar cascade, icon)."""
    if hasattr(sys, '_MEIPASS'):
        bundle_path = os.path.join(sys._MEIPASS, relative_path)
        if os.path.exists(bundle_path):
            return bundle_path
    local_path = os.path.join(BASE_DIR, relative_path)
    if os.path.exists(local_path):
        return local_path
    return relative_path


def load_face_detector():
    """Load Haar Cascade face detector with automatic fallback."""
    cascade_path = get_resource_path(CASCADE_FILE)
    if not os.path.exists(cascade_path) or os.path.getsize(cascade_path) == 0:
        if hasattr(cv2, 'data') and hasattr(cv2.data, 'haarcascades'):
            cascade_path = os.path.join(cv2.data.haarcascades, CASCADE_FILE)
    detector = cv2.CascadeClassifier(cascade_path)
    return detector


def open_camera():
    """DirectShow primary camera loader with MSMF fallback."""
    for idx in [0, 1, 2]:
        for backend in [cv2.CAP_DSHOW, cv2.CAP_MSMF, cv2.CAP_ANY]:
            cap = cv2.VideoCapture(idx, backend)
            if cap.isOpened():
                ret, frame = cap.read()
                if ret and frame is not None:
                    return cap
                cap.release()
    return cv2.VideoCapture(0)


def initialize_storage_files():
    """Ensure students/ directory and both attendance.csv and attendance.xlsx exist with headers."""
    # Ensure students directory exists
    if not os.path.exists(STUDENT_FOLDER):
        try:
            os.makedirs(STUDENT_FOLDER, exist_ok=True)
        except Exception:
            pass

    # 1. CSV
    if not os.path.exists(ATTENDANCE_CSV) or os.path.getsize(ATTENDANCE_CSV) == 0:
        with open(ATTENDANCE_CSV, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Student Name", "Date", "Time", "Status"])

    # 2. XLSX
    if not os.path.exists(ATTENDANCE_XLSX) or os.path.getsize(ATTENDANCE_XLSX) == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Student Name", "Date", "Time", "Status"])
        for col in range(1, 5):
            ws.cell(row=1, column=col).font = Font(bold=True)
        try:
            wb.save(ATTENDANCE_XLSX)
        except Exception:
            pass


def is_already_marked(student_name, today_str):
    """Check if student is already marked present for today in attendance.csv."""
    if os.path.exists(ATTENDANCE_CSV):
        try:
            with open(ATTENDANCE_CSV, mode='r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader, None)  # skip header
                for row in reader:
                    if row and len(row) >= 1 and str(row[0]).strip().lower() == str(student_name).strip().lower():
                        if len(row) >= 4 and row[1] == today_str:
                            return True
                        elif len(row) == 2:  # legacy row fallback
                            return True
        except Exception:
            pass
    return False


def mark_attendance(student_name):
    """Save attendance to both CSV and XLSX seamlessly."""
    initialize_storage_files()
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%I:%M:%S %p")

    if is_already_marked(student_name, today_str):
        return False  # Already marked today

    # 1. Save CSV
    try:
        with open(ATTENDANCE_CSV, mode='a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([student_name, today_str, time_str, "Present"])
    except Exception as e:
        print(f"CSV Save Error: {e}")

    # 2. Save XLSX
    try:
        if os.path.exists(ATTENDANCE_XLSX) and os.path.getsize(ATTENDANCE_XLSX) > 0:
            wb = load_workbook(ATTENDANCE_XLSX)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Student Name", "Date", "Time", "Status"])

        ws.append([student_name, today_str, time_str, "Present"])
        wb.save(ATTENDANCE_XLSX)
    except Exception as e:
        print(f"Excel Save Error: {e}")

    return True


def train_recognizer(face_detector):
    """Train LBPH face recognizer from registered student folders."""
    faces = []
    labels = []
    label_to_name = {}
    name_to_label = {}
    current_label = 0

    if not os.path.exists(STUDENT_FOLDER):
        return None, {}

    # Scan student subdirectories
    dirs = [d for d in os.listdir(STUDENT_FOLDER) if os.path.isdir(os.path.join(STUDENT_FOLDER, d))]
    for student_name in sorted(dirs):
        student_dir = os.path.join(STUDENT_FOLDER, student_name)
        for img_name in sorted(os.listdir(student_dir)):
            img_path = os.path.join(student_dir, img_name)
            if not os.path.isfile(img_path):
                continue
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if img is None or img.size == 0:
                continue

            img_resized = cv2.resize(img, FACE_SIZE)
            img_equalized = cv2.equalizeHist(img_resized)

            if student_name not in name_to_label:
                name_to_label[student_name] = current_label
                label_to_name[current_label] = student_name
                current_label += 1

            faces.append(img_equalized)
            labels.append(name_to_label[student_name])

    # Scan loose images fallback
    files = [f for f in os.listdir(STUDENT_FOLDER)
             if os.path.isfile(os.path.join(STUDENT_FOLDER, f)) and f.lower().endswith(('.jpg', '.jpeg', '.png'))]
    for file_name in sorted(files):
        student_name = os.path.splitext(file_name)[0]
        if student_name in name_to_label:
            continue
        file_path = os.path.join(STUDENT_FOLDER, file_name)
        img = cv2.imread(file_path, cv2.IMREAD_GRAYSCALE)
        if img is None or img.size == 0:
            continue

        if face_detector is not None:
            detected_faces = face_detector.detectMultiScale(img, scaleFactor=1.1, minNeighbors=4, minSize=(50, 50))
            if len(detected_faces) > 0:
                (fx, fy, fw, fh) = detected_faces[0]
                img = img[fy:fy + fh, fx:fx + fw]

        if img.size > 0:
            img_resized = cv2.resize(img, FACE_SIZE)
            img_equalized = cv2.equalizeHist(img_resized)
            name_to_label[student_name] = current_label
            label_to_name[current_label] = student_name
            current_label += 1
            faces.append(img_equalized)
            labels.append(name_to_label[student_name])

    if len(faces) == 0:
        return None, {}

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.train(faces, np.array(labels))
    return recognizer, label_to_name


# ============================================================
# HELPER: INTERACTIVE BUTTON HOVER EFFECTS
# ============================================================

def bind_hover(btn, normal_bg, hover_bg, normal_fg=None, hover_fg=None):
    """Attach smooth hover visual transition to Tkinter buttons."""
    def on_enter(event):
        if btn['state'] != 'disabled':
            btn.configure(bg=hover_bg)
            if hover_fg:
                btn.configure(fg=hover_fg)

    def on_leave(event):
        if btn['state'] != 'disabled':
            btn.configure(bg=normal_bg)
            if normal_fg:
                btn.configure(fg=normal_fg)

    btn.bind("<Enter>", on_enter, add="+")
    btn.bind("<Leave>", on_leave, add="+")


# ============================================================
# MODERN PROFESSIONAL GUI (SANJIVANI ATTENDANCE)
# ============================================================

class SanjivaniAttendanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("SANJIVANI ATTENDANCE — AI-Powered Face Recognition & Smart Attendance Management")
        self.root.geometry("1260x780")
        self.root.minsize(1100, 700)

        # Set Window & Taskbar Icon (Multi-Resolution)
        icon_path = get_resource_path("sanjivani.ico")
        if not os.path.exists(icon_path):
            icon_path = get_resource_path("icon.ico")
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(default=icon_path)
            except Exception:
                try:
                    self.root.iconbitmap(icon_path)
                except Exception:
                    pass

        png_path = get_resource_path("sanjivani.png")
        if not os.path.exists(png_path):
            png_path = get_resource_path("icon.png")
        if os.path.exists(png_path):
            try:
                from PIL import ImageTk, Image
                self._app_icon_photo = ImageTk.PhotoImage(Image.open(png_path))
                self.root.iconphoto(True, self._app_icon_photo)
            except Exception:
                try:
                    self._app_icon_photo = tk.PhotoImage(file=png_path)
                    self.root.iconphoto(True, self._app_icon_photo)
                except Exception:
                    pass

        # ----------------------------------------------------
        # REFINED PROFESSIONAL COLOR PALETTE
        # ----------------------------------------------------
        self.COLOR_PRIMARY = "#0F172A"       # Deep Navy Slate (Header / Base)
        self.COLOR_SECONDARY = "#1E293B"     # Slate 800 (Navbar / Dark Cards)
        self.COLOR_ACCENT = "#2563EB"        # Royal Blue Accent
        self.COLOR_ACCENT_HOVER = "#1D4ED8"  # Deep Blue Hover
        self.COLOR_ACCENT_LIGHT = "#EFF6FF"  # Soft Blue 50 Tint
        self.COLOR_ACCENT_CYAN = "#38BDF8"   # Sky Cyan Highlight
        self.COLOR_MAIN_BG = "#F8FAFC"       # Clean Slate 50 Background
        self.COLOR_CARD_BG = "#FFFFFF"       # Crisp White Card Surface
        self.COLOR_BORDER = "#E2E8F0"        # Subtle Slate 200 Border
        self.COLOR_BORDER_STRONG = "#CBD5E1" # Slate 300 Focus Border

        # Status Tones
        self.COLOR_SUCCESS = "#16A34A"       # Green 600
        self.COLOR_SUCCESS_HOVER = "#15803D" # Green 700
        self.COLOR_SUCCESS_BG = "#DCFCE7"    # Green 100 Soft Tint
        self.COLOR_SUCCESS_TEXT = "#15803D"  # Green 700 Dark Text

        self.COLOR_WARNING = "#F59E0B"       # Amber 500
        self.COLOR_WARNING_BG = "#FEF3C7"    # Amber 100 Soft Tint
        self.COLOR_WARNING_TEXT = "#B45309"  # Amber 700 Dark Text

        self.COLOR_ERROR = "#DC2626"         # Red 600
        self.COLOR_ERROR_HOVER = "#B91C1C"   # Red 700
        self.COLOR_ERROR_BG = "#FEE2E2"      # Red 100 Soft Tint
        self.COLOR_ERROR_TEXT = "#991B1B"    # Red 800 Dark Text

        # Text Hierarchy
        self.TEXT_PRIMARY = "#0F172A"        # Slate 900
        self.TEXT_SECONDARY = "#64748B"      # Slate 500
        self.TEXT_MUTED = "#94A3B8"          # Slate 400
        self.TEXT_WHITE = "#FFFFFF"

        # Typography Family
        self.FONT_FAMILY = "Segoe UI"

        # Apply root background
        self.root.configure(bg=self.COLOR_MAIN_BG)

        self.active_tab = "dashboard"
        self.is_registration_unlocked = False
        self.SECURITY_PIN = "7754"

        # Recognition State Tracking
        self.last_recognized_student = None
        self.last_recognized_confidence = None
        self.last_recognized_time = None
        self.last_recognized_status = None

        self.setup_styles()
        self.create_header()
        self.create_nav_tabs()

        # Main Workspace Container Frame
        self.container = tk.Frame(self.root, bg=self.COLOR_MAIN_BG)
        self.container.pack(fill="both", expand=True, padx=24, pady=(14, 8))

        # Pages
        self.page_dashboard = tk.Frame(self.container, bg=self.COLOR_MAIN_BG)
        self.page_register = tk.Frame(self.container, bg=self.COLOR_MAIN_BG)

        self.build_dashboard_page()
        self.build_register_page()

        # Status & Notification Bar
        self.create_status_bar()

        # Show initial page
        self.show_page("dashboard")

        # Start live clock timer
        self.update_live_clock()

    def setup_styles(self):
        """Configure clean modern TTK Treeview and Progressbar styles."""
        style = ttk.Style()
        style.theme_use("clam")

        # Configure Treeview Header
        style.configure(
            "Custom.Treeview.Heading",
            font=(self.FONT_FAMILY, 9, "bold"),
            background="#F1F5F9",
            foreground=self.COLOR_SECONDARY,
            relief="flat",
            padding=9,
            borderwidth=0
        )
        style.map(
            "Custom.Treeview.Heading",
            background=[("active", "#E2E8F0")]
        )

        # Configure Treeview Rows
        style.configure(
            "Custom.Treeview",
            font=(self.FONT_FAMILY, 9),
            rowheight=36,
            background=self.COLOR_CARD_BG,
            fieldbackground=self.COLOR_CARD_BG,
            foreground=self.TEXT_PRIMARY,
            borderwidth=0,
            relief="flat"
        )
        style.map(
            "Custom.Treeview",
            background=[("selected", self.COLOR_ACCENT_LIGHT)],
            foreground=[("selected", self.COLOR_ACCENT)]
        )

        # Progressbar styling
        style.configure(
            "Modern.Horizontal.TProgressbar",
            troughcolor=self.COLOR_BORDER,
            background=self.COLOR_ACCENT,
            lightcolor=self.COLOR_ACCENT_CYAN,
            darkcolor=self.COLOR_ACCENT_HOVER,
            bordercolor=self.COLOR_BORDER,
            thickness=10
        )

        # Scrollbar styling
        style.configure(
            "Vertical.TScrollbar",
            gripcount=0,
            background="#CBD5E1",
            darkcolor="#94A3B8",
            lightcolor="#E2E8F0",
            troughcolor=self.COLOR_MAIN_BG,
            bordercolor=self.COLOR_MAIN_BG,
            arrowcolor=self.COLOR_SECONDARY
        )

    # ============================================================
    # 1. HEADER (Executive Navy & Modern Status Display)
    # ============================================================

    def create_header(self):
        header = tk.Frame(self.root, bg=self.COLOR_PRIMARY, height=76)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        # Left Branding Block
        brand_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        brand_frame.pack(side="left", padx=26, pady=12)

        # Title Row + AI Pill Badge
        title_line = tk.Frame(brand_frame, bg=self.COLOR_PRIMARY)
        title_line.pack(anchor="w")

        main_title = tk.Label(
            title_line,
            text="SANJIVANI ATTENDANCE",
            font=(self.FONT_FAMILY, 16, "bold"),
            fg=self.TEXT_WHITE,
            bg=self.COLOR_PRIMARY
        )
        main_title.pack(side="left")

        ai_badge = tk.Label(
            title_line,
            text="  ⚡ AI ATTENDANCE SYSTEM  ",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.COLOR_ACCENT_CYAN,
            bg=self.COLOR_SECONDARY,
            bd=1,
            relief="solid",
            highlightbackground="#334155",
            padx=6,
            pady=2
        )
        ai_badge.pack(side="left", padx=(12, 0))

        sub_title = tk.Label(
            brand_frame,
            text="Enterprise AI Face Recognition & Smart Attendance Management",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_MUTED,
            bg=self.COLOR_PRIMARY
        )
        sub_title.pack(anchor="w", pady=(3, 0))

        # Right Live Clock & System Status Indicator
        clock_frame = tk.Frame(header, bg=self.COLOR_PRIMARY)
        clock_frame.pack(side="right", padx=26, pady=10)

        # System Online Pill
        status_row = tk.Frame(clock_frame, bg=self.COLOR_PRIMARY)
        status_row.pack(anchor="e", pady=(0, 3))

        self.lbl_system_status = tk.Label(
            status_row,
            text="● SYSTEM ONLINE",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.COLOR_SUCCESS,
            bg="#064E3B",
            padx=10,
            pady=2
        )
        self.lbl_system_status.pack(side="right")

        # Time & Date Row
        datetime_row = tk.Frame(clock_frame, bg=self.COLOR_PRIMARY)
        datetime_row.pack(anchor="e")

        self.lbl_clock_time = tk.Label(
            datetime_row,
            text="--:--:-- PM",
            font=(self.FONT_FAMILY, 13, "bold"),
            fg=self.COLOR_ACCENT_CYAN,
            bg=self.COLOR_PRIMARY
        )
        self.lbl_clock_time.pack(side="left")

        tk.Label(
            datetime_row,
            text="  |  ",
            font=(self.FONT_FAMILY, 10),
            fg="#475569",
            bg=self.COLOR_PRIMARY
        ).pack(side="left")

        self.lbl_clock_date = tk.Label(
            datetime_row,
            text=datetime.now().strftime("%A, %b %d, %Y"),
            font=(self.FONT_FAMILY, 9),
            fg="#CBD5E1",
            bg=self.COLOR_PRIMARY
        )
        self.lbl_clock_date.pack(side="left")

        # Bottom Accent Underline (2px Royal Blue)
        accent_line = tk.Frame(self.root, bg=self.COLOR_ACCENT, height=2)
        accent_line.pack(fill="x", side="top")

    def update_live_clock(self):
        now = datetime.now()
        self.lbl_clock_time.configure(text=now.strftime("%I:%M:%S %p"))
        self.lbl_clock_date.configure(text=now.strftime("%A, %b %d, %Y"))
        self.root.after(1000, self.update_live_clock)

    # ============================================================
    # 2. NAVIGATION (Tabs)
    # ============================================================

    def create_nav_tabs(self):
        nav = tk.Frame(self.root, bg=self.COLOR_SECONDARY, height=46)
        nav.pack(fill="x", side="top")
        nav.pack_propagate(False)

        nav_inner = tk.Frame(nav, bg=self.COLOR_SECONDARY)
        nav_inner.pack(side="left", padx=22, pady=5)

        self.btn_tab_dash = tk.Button(
            nav_inner,
            text="  📊  Dashboard  ",
            font=(self.FONT_FAMILY, 9, "bold"),
            bg=self.COLOR_ACCENT,
            fg=self.TEXT_WHITE,
            bd=0,
            activebackground=self.COLOR_ACCENT_HOVER,
            activeforeground=self.TEXT_WHITE,
            cursor="hand2",
            padx=16,
            pady=5,
            command=lambda: self.show_page("dashboard")
        )
        self.btn_tab_dash.pack(side="left", padx=(0, 8))

        self.btn_tab_reg = tk.Button(
            nav_inner,
            text="  👤  Student Registration  ",
            font=(self.FONT_FAMILY, 9, "bold"),
            bg=self.COLOR_SECONDARY,
            fg=self.TEXT_MUTED,
            bd=0,
            activebackground="#334155",
            activeforeground=self.TEXT_WHITE,
            cursor="hand2",
            padx=16,
            pady=5,
            command=lambda: self.show_page("register")
        )
        self.btn_tab_reg.pack(side="left", padx=0)

        # Tab hover handlers
        self.setup_tab_hovers()

    def setup_tab_hovers(self):
        def on_enter_dash(e):
            if self.active_tab != "dashboard":
                self.btn_tab_dash.configure(bg="#334155", fg=self.TEXT_WHITE)
        def on_leave_dash(e):
            if self.active_tab != "dashboard":
                self.btn_tab_dash.configure(bg=self.COLOR_SECONDARY, fg=self.TEXT_MUTED)

        def on_enter_reg(e):
            if self.active_tab != "register":
                self.btn_tab_reg.configure(bg="#334155", fg=self.TEXT_WHITE)
        def on_leave_reg(e):
            if self.active_tab != "register":
                self.btn_tab_reg.configure(bg=self.COLOR_SECONDARY, fg=self.TEXT_MUTED)

        self.btn_tab_dash.bind("<Enter>", on_enter_dash)
        self.btn_tab_dash.bind("<Leave>", on_leave_dash)
        self.btn_tab_reg.bind("<Enter>", on_enter_reg)
        self.btn_tab_reg.bind("<Leave>", on_leave_reg)

    def prompt_security_lock(self):
        """Prompt admin PIN dialog (7754) before opening Registration page."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Security Lock — Admin Authentication")
        dialog.geometry("390x260")
        dialog.resizable(False, False)
        dialog.configure(bg=self.COLOR_PRIMARY)
        dialog.transient(self.root)
        dialog.grab_set()

        # Center dialog relative to root
        self.root.update_idletasks()
        rx = self.root.winfo_x()
        ry = self.root.winfo_y()
        rw = self.root.winfo_width()
        rh = self.root.winfo_height()
        dialog.geometry(f"+{rx + (rw // 2) - 195}+{ry + (rh // 2) - 130}")

        tk.Label(
            dialog,
            text="🔒 ADMIN SECURITY LOCK",
            font=(self.FONT_FAMILY, 12, "bold"),
            fg=self.TEXT_WHITE,
            bg=self.COLOR_PRIMARY
        ).pack(pady=(24, 4))

        tk.Label(
            dialog,
            text="Enter 4-Digit Passcode to Unlock Student Registration:",
            font=(self.FONT_FAMILY, 9),
            fg="#CBD5E1",
            bg=self.COLOR_PRIMARY
        ).pack(pady=(0, 14))

        pin_entry = tk.Entry(
            dialog,
            font=(self.FONT_FAMILY, 14, "bold"),
            show="•",
            justify="center",
            bd=1,
            relief="solid",
            bg=self.COLOR_CARD_BG,
            fg=self.TEXT_PRIMARY,
            highlightthickness=1,
            highlightbackground=self.COLOR_ACCENT,
            width=12
        )
        pin_entry.pack(pady=(0, 20), ipady=4)
        pin_entry.focus_set()

        success = [False]

        def check_pin():
            if pin_entry.get().strip() == self.SECURITY_PIN:
                success[0] = True
                dialog.destroy()
            else:
                messagebox.showerror("Access Denied", "❌ Incorrect Security Passcode! Access Denied.", parent=dialog)
                pin_entry.delete(0, tk.END)

        btn_frame = tk.Frame(dialog, bg=self.COLOR_PRIMARY)
        btn_frame.pack(fill="x", padx=40)

        btn_unlock = tk.Button(
            btn_frame,
            text=" Unlock Page ",
            font=(self.FONT_FAMILY, 9, "bold"),
            bg=self.COLOR_ACCENT,
            fg=self.TEXT_WHITE,
            activebackground=self.COLOR_ACCENT_HOVER,
            activeforeground=self.TEXT_WHITE,
            bd=0,
            cursor="hand2",
            command=check_pin,
            padx=12,
            pady=7
        )
        btn_unlock.pack(side="left", expand=True, fill="x", padx=(0, 6))
        bind_hover(btn_unlock, self.COLOR_ACCENT, self.COLOR_ACCENT_HOVER)

        btn_cancel = tk.Button(
            btn_frame,
            text=" Cancel ",
            font=(self.FONT_FAMILY, 9),
            bg="#334155",
            fg=self.TEXT_WHITE,
            activebackground="#475569",
            activeforeground=self.TEXT_WHITE,
            bd=0,
            cursor="hand2",
            command=dialog.destroy,
            padx=12,
            pady=7
        )
        btn_cancel.pack(side="right", expand=True, fill="x", padx=(6, 0))
        bind_hover(btn_cancel, "#334155", "#475569")

        dialog.bind("<Return>", lambda e: check_pin())
        self.root.wait_window(dialog)
        return success[0]

    def show_page(self, page_name):
        if page_name == "dashboard":
            self.active_tab = "dashboard"
            self.page_register.pack_forget()
            self.page_dashboard.pack(fill="both", expand=True)

            self.btn_tab_dash.configure(bg=self.COLOR_ACCENT, fg=self.TEXT_WHITE)
            self.btn_tab_reg.configure(bg=self.COLOR_SECONDARY, fg=self.TEXT_MUTED)
            self.refresh_dashboard()

        elif page_name == "register":
            if not self.is_registration_unlocked:
                unlocked = self.prompt_security_lock()
                if not unlocked:
                    self.set_status("❌ Access Denied — Registration page locked.")
                    return
                self.is_registration_unlocked = True
                self.set_status("✓ Access Granted — Student Registration Unlocked.")

            self.active_tab = "register"
            self.page_dashboard.pack_forget()
            self.page_register.pack(fill="both", expand=True)

            self.btn_tab_reg.configure(bg=self.COLOR_ACCENT, fg=self.TEXT_WHITE)
            self.btn_tab_dash.configure(bg=self.COLOR_SECONDARY, fg=self.TEXT_MUTED)
            self.refresh_students_directory()

    # ============================================================
    # 3. DASHBOARD PAGE (Analytics, Direct Face Attendance, Table)
    # ============================================================

    def build_dashboard_page(self):
        # ----------------------------------------------------
        # TOP ROW: 4 Balanced Stat Cards
        # ----------------------------------------------------
        stats_grid = tk.Frame(self.page_dashboard, bg=self.COLOR_MAIN_BG)
        stats_grid.pack(fill="x", pady=(0, 14))

        c1 = self.create_stat_card(
            stats_grid,
            title="TOTAL REGISTERED",
            default_val="0",
            subtext="👥 Enrolled in System",
            val_color=self.TEXT_PRIMARY,
            accent_color=self.COLOR_ACCENT
        )
        c1.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.lbl_stat_students = c1.lbl_value

        c2 = self.create_stat_card(
            stats_grid,
            title="PRESENT TODAY",
            default_val="0",
            subtext="🟢 Checked In Today",
            val_color=self.COLOR_SUCCESS,
            accent_color=self.COLOR_SUCCESS
        )
        c2.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.lbl_stat_present = c2.lbl_value

        c3 = self.create_stat_card(
            stats_grid,
            title="ABSENT TODAY",
            default_val="0",
            subtext="🔴 Pending Attendance",
            val_color=self.COLOR_ERROR,
            accent_color=self.COLOR_ERROR
        )
        c3.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.lbl_stat_absent = c3.lbl_value

        c4 = self.create_stat_card(
            stats_grid,
            title="ATTENDANCE RATE",
            default_val="0.0%",
            subtext="📈 Daily Turnout Percentage",
            val_color=self.COLOR_ACCENT,
            accent_color="#6366F1"
        )
        c4.pack(side="left", fill="x", expand=True)
        self.lbl_stat_rate = c4.lbl_value

        # ----------------------------------------------------
        # MIDDLE & BOTTOM SPLIT LAYOUT:
        # Left Panel (Width ~380): Direct Face Attendance + Result Card
        # Right Panel (Expand): Attendance Records Table + Toolbar
        # ----------------------------------------------------
        content_split = tk.Frame(self.page_dashboard, bg=self.COLOR_MAIN_BG)
        content_split.pack(fill="both", expand=True)

        # ----------------------------------------------------
        # LEFT PANEL
        # ----------------------------------------------------
        left_panel = tk.Frame(content_split, bg=self.COLOR_MAIN_BG, width=380)
        left_panel.pack(side="left", fill="y", padx=(0, 14))
        left_panel.pack_propagate(False)

        # 1. DIRECT FACE ATTENDANCE CARD
        face_card = tk.LabelFrame(
            left_panel,
            text=" DIRECT FACE ATTENDANCE ",
            font=(self.FONT_FAMILY, 10, "bold"),
            fg=self.TEXT_PRIMARY,
            bg=self.COLOR_CARD_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=16,
            pady=14
        )
        face_card.pack(fill="x", pady=(0, 14))

        tk.Label(
            face_card,
            text="Verify student biometric identity and automatically record daily attendance.",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG,
            wraplength=320,
            justify="left"
        ).pack(anchor="w", pady=(0, 12))

        # Status Badge Indicator Box
        status_box = tk.Frame(
            face_card,
            bg=self.COLOR_MAIN_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=12,
            pady=7
        )
        status_box.pack(fill="x", pady=(0, 14))

        tk.Label(
            status_box,
            text="System Status:",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_MAIN_BG
        ).pack(side="left")

        self.lbl_camera_status = tk.Label(
            status_box,
            text="● Camera Ready",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.COLOR_SUCCESS,
            bg=self.COLOR_MAIN_BG
        )
        self.lbl_camera_status.pack(side="right")

        # Large Primary CTA Button: START FACE ATTENDANCE
        self.btn_face_scan = tk.Button(
            face_card,
            text="  📸  START FACE ATTENDANCE  ",
            font=(self.FONT_FAMILY, 10, "bold"),
            bg=self.COLOR_SUCCESS,
            fg=self.TEXT_WHITE,
            activebackground=self.COLOR_SUCCESS_HOVER,
            activeforeground=self.TEXT_WHITE,
            bd=0,
            cursor="hand2",
            pady=10,
            command=self.start_direct_face_attendance
        )
        self.btn_face_scan.pack(fill="x", pady=(0, 2))
        bind_hover(self.btn_face_scan, self.COLOR_SUCCESS, self.COLOR_SUCCESS_HOVER)

        # 2. RECOGNITION RESULT CARD
        self.result_card = tk.LabelFrame(
            left_panel,
            text=" LATEST RECOGNITION RESULT ",
            font=(self.FONT_FAMILY, 10, "bold"),
            fg=self.TEXT_PRIMARY,
            bg=self.COLOR_CARD_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=16,
            pady=14
        )
        self.result_card.pack(fill="both", expand=True)

        # Verification Status Badge
        self.lbl_result_badge = tk.Label(
            self.result_card,
            text="AWAITING ATTENDANCE SCAN",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.TEXT_SECONDARY,
            bg="#F1F5F9",
            padx=12,
            pady=5
        )
        self.lbl_result_badge.pack(fill="x", pady=(0, 12))

        # Details Key-Value Frame
        details_frame = tk.Frame(self.result_card, bg=self.COLOR_CARD_BG)
        details_frame.pack(fill="x", pady=2)

        # Student Name
        tk.Label(
            details_frame,
            text="Student Name:",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        ).grid(row=0, column=0, sticky="w", pady=5)

        self.lbl_res_name = tk.Label(
            details_frame,
            text="--",
            font=(self.FONT_FAMILY, 10, "bold"),
            fg=self.TEXT_PRIMARY,
            bg=self.COLOR_CARD_BG
        )
        self.lbl_res_name.grid(row=0, column=1, sticky="w", padx=(12, 0), pady=5)

        # Confidence
        tk.Label(
            details_frame,
            text="Confidence:",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        ).grid(row=1, column=0, sticky="w", pady=5)

        self.lbl_res_conf = tk.Label(
            details_frame,
            text="--",
            font=(self.FONT_FAMILY, 9, "bold"),
            fg=self.COLOR_ACCENT,
            bg=self.COLOR_CARD_BG
        )
        self.lbl_res_conf.grid(row=1, column=1, sticky="w", padx=(12, 0), pady=5)

        # Status
        tk.Label(
            details_frame,
            text="Status:",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        ).grid(row=2, column=0, sticky="w", pady=5)

        self.lbl_res_status = tk.Label(
            details_frame,
            text="--",
            font=(self.FONT_FAMILY, 9, "bold"),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        )
        self.lbl_res_status.grid(row=2, column=1, sticky="w", padx=(12, 0), pady=5)

        # Timestamp
        tk.Label(
            details_frame,
            text="Timestamp:",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        ).grid(row=3, column=0, sticky="w", pady=5)

        self.lbl_res_time = tk.Label(
            details_frame,
            text="--",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        )
        self.lbl_res_time.grid(row=3, column=1, sticky="w", padx=(12, 0), pady=5)

        # ----------------------------------------------------
        # RIGHT PANEL: Attendance Records Table & Actions
        # ----------------------------------------------------
        records_card = tk.LabelFrame(
            content_split,
            text=" TODAY'S ATTENDANCE LOG & ACTIONS ",
            font=(self.FONT_FAMILY, 10, "bold"),
            fg=self.TEXT_PRIMARY,
            bg=self.COLOR_CARD_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=16,
            pady=14
        )
        records_card.pack(side="right", fill="both", expand=True)

        # Action Toolbar
        toolbar = tk.Frame(records_card, bg=self.COLOR_CARD_BG)
        toolbar.pack(fill="x", pady=(0, 12))

        # Email Report Button (Primary Blue)
        btn_email = tk.Button(
            toolbar,
            text="  📧  SEND ATTENDANCE REPORT  ",
            font=(self.FONT_FAMILY, 9, "bold"),
            bg=self.COLOR_ACCENT,
            fg=self.TEXT_WHITE,
            activebackground=self.COLOR_ACCENT_HOVER,
            activeforeground=self.TEXT_WHITE,
            bd=0,
            cursor="hand2",
            padx=14,
            pady=7,
            command=self.handle_send_email_report
        )
        btn_email.pack(side="left")
        bind_hover(btn_email, self.COLOR_ACCENT, self.COLOR_ACCENT_HOVER)

        # Refresh Button (Secondary White with subtle border)
        btn_refresh = tk.Button(
            toolbar,
            text=" 🔄 Refresh Table ",
            font=(self.FONT_FAMILY, 9, "bold"),
            bg=self.COLOR_CARD_BG,
            fg=self.TEXT_PRIMARY,
            activebackground="#F1F5F9",
            activeforeground=self.TEXT_PRIMARY,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            cursor="hand2",
            padx=12,
            pady=6,
            command=self.refresh_dashboard
        )
        btn_refresh.pack(side="right", padx=(8, 0))
        bind_hover(btn_refresh, self.COLOR_CARD_BG, "#F1F5F9")

        # Reset Log Button (Danger Soft Red)
        btn_clear = tk.Button(
            toolbar,
            text=" 🗑 Reset Log ",
            font=(self.FONT_FAMILY, 9, "bold"),
            bg=self.COLOR_ERROR_BG,
            fg=self.COLOR_ERROR,
            activebackground="#FECACA",
            activeforeground=self.COLOR_ERROR_HOVER,
            bd=1,
            relief="solid",
            highlightbackground="#FECACA",
            cursor="hand2",
            padx=12,
            pady=6,
            command=self.clear_attendance_records
        )
        btn_clear.pack(side="right")
        bind_hover(btn_clear, self.COLOR_ERROR_BG, "#FEE2E2")

        # Attendance Table Frame
        table_container = tk.Frame(records_card, bg=self.COLOR_CARD_BG, bd=1, relief="solid", highlightbackground=self.COLOR_BORDER)
        table_container.pack(fill="both", expand=True)

        columns = ("name", "date", "time", "status", "confidence")
        self.dash_tree = ttk.Treeview(
            table_container,
            columns=columns,
            show="headings",
            selectmode="browse",
            style="Custom.Treeview"
        )

        self.dash_tree.heading("name", text="Student Name")
        self.dash_tree.heading("date", text="Date")
        self.dash_tree.heading("time", text="Time Marked")
        self.dash_tree.heading("status", text="Status")
        self.dash_tree.heading("confidence", text="Confidence / Match")

        self.dash_tree.column("name", width=190, anchor="w")
        self.dash_tree.column("date", width=110, anchor="center")
        self.dash_tree.column("time", width=120, anchor="center")
        self.dash_tree.column("status", width=110, anchor="center")
        self.dash_tree.column("confidence", width=140, anchor="center")

        self.dash_tree.tag_configure("present_badge", foreground=self.COLOR_SUCCESS, font=(self.FONT_FAMILY, 9, "bold"))
        self.dash_tree.tag_configure("absent_badge", foreground=self.COLOR_ERROR, font=(self.FONT_FAMILY, 9, "bold"))
        self.dash_tree.tag_configure("empty_msg", foreground=self.TEXT_MUTED, font=(self.FONT_FAMILY, 9, "italic"))

        scrollbar = ttk.Scrollbar(table_container, orient="vertical", command=self.dash_tree.yview, style="Vertical.TScrollbar")
        self.dash_tree.configure(yscrollcommand=scrollbar.set)

        self.dash_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def create_stat_card(self, parent, title, default_val, subtext, val_color, accent_color):
        card = tk.Frame(parent, bg=self.COLOR_CARD_BG, bd=1, relief="solid", highlightbackground=self.COLOR_BORDER)
        card.pack_propagate(False)

        # Top subtle colored accent stripe (3px)
        tk.Frame(card, bg=accent_color, height=3).pack(fill="x", side="top")

        content = tk.Frame(card, bg=self.COLOR_CARD_BG)
        content.pack(fill="both", expand=True, padx=14, pady=(8, 10))

        tk.Label(
            content,
            text=title,
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        ).pack(anchor="w")

        lbl_val = tk.Label(
            content,
            text=default_val,
            font=(self.FONT_FAMILY, 22, "bold"),
            fg=val_color,
            bg=self.COLOR_CARD_BG
        )
        lbl_val.pack(anchor="w", pady=(2, 2))

        tk.Label(
            content,
            text=subtext,
            font=(self.FONT_FAMILY, 8),
            fg=self.TEXT_MUTED,
            bg=self.COLOR_CARD_BG
        ).pack(anchor="w")

        card.lbl_value = lbl_val
        return card

    # ============================================================
    # 4. DIRECT FACE RECOGNITION ATTENDANCE (PRESERVED BACKEND)
    # ============================================================

    def start_direct_face_attendance(self):
        """AI Face Recognition Attendance System."""
        detector = load_face_detector()
        self.set_camera_status("● Recognizing...", self.COLOR_ACCENT)
        self.set_status("ℹ Training face recognizer model from registered faces...")
        self.root.update_idletasks()

        recognizer, label_map = train_recognizer(detector)

        if recognizer is None:
            self.set_camera_status("● No Data", self.COLOR_ERROR)
            messagebox.showwarning(
                "No Student Data",
                "No registered student faces found!\nPlease register students first from the Student Registration tab."
            )
            self.set_status("⚠ Attendance cancelled — no registered face data.")
            self.set_camera_status("● Camera Ready", self.COLOR_SUCCESS)
            return

        camera = open_camera()
        if not camera.isOpened():
            self.set_camera_status("● Camera Error", self.COLOR_ERROR)
            messagebox.showerror(
                "Camera Error",
                "Could not open camera device! Please check webcam connection."
            )
            self.set_status("✕ Camera could not be opened.")
            return

        self.set_camera_status("● Camera Active", self.COLOR_SUCCESS)
        self.set_status("ℹ SANJIVANI AI Face Recognition Active — Look into camera...")
        stop_program = False

        while True:
            success, frame = camera.read()
            if not success:
                break

            h_f, w_f = frame.shape[:2]
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(100, 100))

            # Modern HUD Header in Camera Window (Midnight Navy)
            cv2.rectangle(frame, (0, 0), (w_f, 48), (42, 23, 15), -1)
            cv2.putText(frame, "SANJIVANI ATTENDANCE — AI FACE RECOGNITION", (15, 32),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.65, (248, 189, 56), 2)

            if len(faces) == 0:
                cv2.rectangle(frame, (0, h_f - 32), (w_f, h_f), (42, 23, 15), -1)
                cv2.putText(frame, "LIVE CAMERA: Looking for face... | Press 'Q' to Exit", (15, h_f - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (184, 163, 148), 1)

            for (x, y, w, h) in faces:
                if w <= 0 or h <= 0:
                    continue
                face_roi = gray[y:y + h, x:x + w]
                if face_roi is None or face_roi.size == 0:
                    continue

                face_resized = cv2.resize(face_roi, FACE_SIZE)
                face_equalized = cv2.equalizeHist(face_resized)
                label, confidence = recognizer.predict(face_equalized)

                # Confidence calculation for display (Lower distance = higher match confidence)
                match_pct = max(0.0, min(100.0, 100.0 - (confidence * 0.7)))

                if confidence < CONFIDENCE_THRESHOLD:
                    student_name = label_map.get(label, "Unknown")
                    was_marked = mark_attendance(student_name)
                    now_str = datetime.now().strftime("%I:%M:%S %p")
                    date_str = datetime.now().strftime("%Y-%m-%d")

                    if was_marked:
                        disp_text = f"PRESENT: {student_name} ({match_pct:.0f}%)"
                        self.set_status(f"✓ Attendance marked successfully for {student_name}")
                        self.set_camera_status("● Attendance Marked", self.COLOR_SUCCESS)
                        self.update_result_card(
                            verified=True,
                            name=student_name,
                            confidence=f"{match_pct:.1f}%",
                            status="PRESENT",
                            time_str=f"{date_str} {now_str}"
                        )
                        message_title = "✓ Attendance Marked"
                        message_body = f"Attendance marked successfully for {student_name}!\n\nConfidence: {match_pct:.1f}%\nStatus: PRESENT"
                    else:
                        disp_text = f"{student_name} (Already Marked Today)"
                        self.set_status(f"ℹ {student_name} is already marked present today.")
                        self.set_camera_status("● Already Marked", self.COLOR_WARNING)
                        self.update_result_card(
                            verified=True,
                            name=student_name,
                            confidence=f"{match_pct:.1f}%",
                            status="ALREADY MARKED",
                            time_str=f"{date_str} {now_str}"
                        )
                        message_title = "Already Marked Today"
                        message_body = f"{student_name} is already marked present for today."

                    color = (74, 163, 22)  # Emerald Green (BGR)
                    cv2.rectangle(frame, (x, y), (x + w, y + h), color, 3)

                    # Camera HUD Overlay
                    cv2.rectangle(frame, (0, 0), (w_f, 48), (42, 23, 15), -1)
                    cv2.putText(frame, f"SANJIVANI ATTENDANCE - {disp_text}", (15, 32),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.65, color, 2)

                    cv2.rectangle(frame, (0, h_f - 32), (w_f, h_f), (42, 23, 15), -1)
                    cv2.putText(frame, "VERIFIED: Attendance recorded successfully!", (15, h_f - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)

                    cv2.imshow("SANJIVANI ATTENDANCE — Live Recognition", frame)
                    cv2.waitKey(2000)
                    stop_program = True

                    messagebox.showinfo(message_title, message_body)
                    break
                else:
                    # STRICT UNKNOWN FACE REJECTION
                    color = (38, 38, 220)  # Crimson Red (BGR)
                    disp_text = f"UNKNOWN FACE REJECTED ({match_pct:.0f}%)"
                    cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
                    cv2.rectangle(frame, (0, 0), (w_f, 48), (42, 23, 15), -1)
                    cv2.putText(frame, "SANJIVANI ATTENDANCE — UNKNOWN FACE REJECTED", (15, 32),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.65, color, 2)
                    cv2.putText(frame, disp_text, (x, max(y - 10, 25)),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

                    self.set_camera_status("● Face Rejected", self.COLOR_ERROR)
                    self.set_status("⚠ Face not recognized — UNKNOWN FACE REJECTED (Access Denied).")
                    self.update_result_card(
                        verified=False,
                        name="Unknown Individual",
                        confidence=f"{match_pct:.1f}%",
                        status="UNKNOWN (REJECTED)",
                        time_str=datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
                    )

            cv2.imshow("SANJIVANI ATTENDANCE — Live Recognition", frame)
            if stop_program or (cv2.waitKey(1) & 0xFF == ord('q')):
                break

        camera.release()
        cv2.destroyAllWindows()
        self.set_camera_status("● Camera Ready", self.COLOR_SUCCESS)
        self.refresh_dashboard()

    def set_camera_status(self, text, color):
        self.lbl_camera_status.configure(text=text, fg=color)

    def update_result_card(self, verified, name, confidence, status, time_str):
        if verified:
            self.lbl_result_badge.configure(
                text="✓ IDENTITY VERIFIED",
                fg=self.COLOR_SUCCESS_TEXT,
                bg=self.COLOR_SUCCESS_BG
            )
            self.lbl_res_name.configure(text=name, fg=self.TEXT_PRIMARY)
            self.lbl_res_conf.configure(text=confidence, fg=self.COLOR_SUCCESS)
            self.lbl_res_status.configure(text=status, fg=self.COLOR_SUCCESS)
        else:
            self.lbl_result_badge.configure(
                text="✕ IDENTITY NOT VERIFIED",
                fg=self.COLOR_ERROR_TEXT,
                bg=self.COLOR_ERROR_BG
            )
            self.lbl_res_name.configure(text=name, fg=self.COLOR_ERROR)
            self.lbl_res_conf.configure(text=confidence, fg=self.COLOR_ERROR)
            self.lbl_res_status.configure(text=status, fg=self.COLOR_ERROR)

        self.lbl_res_time.configure(text=time_str)

    # ============================================================
    # 5. DASHBOARD DATA REFRESH & ANALYTICS
    # ============================================================

    def refresh_dashboard(self):
        """Reload attendance entries and calculate dashboard analytics."""
        for item in self.dash_tree.get_children():
            self.dash_tree.delete(item)

        initialize_storage_files()

        present_students = []
        today_str = datetime.now().strftime("%Y-%m-%d")

        if os.path.exists(ATTENDANCE_CSV):
            try:
                with open(ATTENDANCE_CSV, mode="r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader, None)  # skip header
                    for row in reader:
                        if not row:
                            continue
                        if len(row) >= 4:
                            name, date_val, time_val, status_val = row[0], row[1], row[2], row[3]
                        elif len(row) == 2:
                            name, date_val, time_val, status_val = row[0], today_str, row[1], "Present"
                        else:
                            continue

                        conf_str = "Verified (High)"
                        self.dash_tree.insert(
                            "", "end",
                            values=(name, date_val, time_val, status_val, conf_str),
                            tags=("present_badge",)
                        )
                        if name not in present_students and date_val == today_str:
                            present_students.append(name)
            except Exception as e:
                self.set_status(f"CSV Reading Warning: {e}")

        # Empty State Handling
        if len(self.dash_tree.get_children()) == 0:
            self.dash_tree.insert(
                "", "end",
                values=("No attendance records logged today yet", "-", "-", "-", "-"),
                tags=("empty_msg",)
            )

        # Total registered students
        reg_count = 0
        if os.path.exists(STUDENT_FOLDER):
            reg_count = len([d for d in os.listdir(STUDENT_FOLDER) if os.path.isdir(os.path.join(STUDENT_FOLDER, d))])

        present_count = len(present_students)
        absent_count = max(0, reg_count - present_count)
        rate = (present_count / reg_count * 100) if reg_count > 0 else 0.0

        # Update Analytics Cards
        self.lbl_stat_students.configure(text=str(reg_count))
        self.lbl_stat_present.configure(text=str(present_count))
        self.lbl_stat_absent.configure(text=str(absent_count))
        self.lbl_stat_rate.configure(text=f"{rate:.1f}%")

        self.set_status(f"✓ Dashboard Refreshed — {present_count} Present, {absent_count} Absent out of {reg_count} Students.")

    def clear_attendance_records(self):
        if messagebox.askyesno(
            "Reset Records & Send Email Report",
            f"Clearing logs will automatically send the complete attendance report and attendance.xlsx to {send_email.RECIPIENT_EMAIL} before resetting.\n\nDo you want to proceed?"
        ):
            self.set_status(f"Sending email report to {send_email.RECIPIENT_EMAIL} before clearing logs...")
            self.root.update_idletasks()

            env_email, env_pass = send_email.get_email_credentials()
            sender_email = env_email
            sender_password = env_pass

            if not sender_email or not sender_password:
                sender_email, sender_password = self.prompt_email_credentials()

            if sender_email and sender_password:
                success, msg = send_email.send_email_report(sender_email, sender_password)
                if success:
                    messagebox.showinfo("Email Sent", f"✓ Attendance report successfully emailed to {send_email.RECIPIENT_EMAIL} before clearing!")
                else:
                    messagebox.showwarning("Email Warning", f"Could not email report before clearing:\n{msg}")

            try:
                # Reset CSV
                with open(ATTENDANCE_CSV, mode='w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Student Name", "Date", "Time", "Status"])

                # Reset XLSX
                wb = Workbook()
                ws = wb.active
                ws.title = "Attendance"
                ws.append(["Student Name", "Date", "Time", "Status"])
                for col in range(1, 5):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                wb.save(ATTENDANCE_XLSX)

                self.refresh_dashboard()
                self.set_status("✓ Attendance logs reset successfully.")
                messagebox.showinfo("Log Reset", "✓ Attendance records emailed and reset successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Could not clear logs: {e}")

    # ============================================================
    # 6. EMAIL REPORT DISPATCH (PRESERVED)
    # ============================================================

    def handle_send_email_report(self):
        """Send email report to rohitvinchu7754@gmail.com with attendance.xlsx attached."""
        env_email, env_pass = send_email.get_email_credentials()
        sender_email = env_email
        sender_password = env_pass

        if not sender_email or not sender_password:
            sender_email, sender_password = self.prompt_email_credentials()
            if not sender_email or not sender_password:
                self.set_status("⚠ Email report cancelled — credentials missing.")
                return

        self.set_status(f"Sending email report to {send_email.RECIPIENT_EMAIL}...")
        self.root.update_idletasks()

        success, message = send_email.send_email_report(sender_email, sender_password)

        if success:
            self.set_status("✓ Attendance report sent successfully.")
            messagebox.showinfo("Email Sent Successfully", f"{message}")
        else:
            self.set_status("⚠ Unable to send attendance report.")
            messagebox.showerror("Email Error", f"Unable to send attendance report:\n\n{message}")

    def prompt_email_credentials(self):
        """Modal dialog to enter Gmail Sender Email & App Password."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Configure Email Credentials")
        dialog.geometry("450x320")
        dialog.resizable(False, False)
        dialog.configure(bg=self.COLOR_PRIMARY)
        dialog.transient(self.root)
        dialog.grab_set()

        self.root.update_idletasks()
        rx, ry, rw, rh = self.root.winfo_x(), self.root.winfo_y(), self.root.winfo_width(), self.root.winfo_height()
        dialog.geometry(f"+{rx + (rw // 2) - 225}+{ry + (rh // 2) - 160}")

        tk.Label(
            dialog,
            text="📧 SENDER EMAIL CONFIGURATION",
            font=(self.FONT_FAMILY, 11, "bold"),
            fg=self.TEXT_WHITE,
            bg=self.COLOR_PRIMARY
        ).pack(pady=(20, 4))

        tk.Label(
            dialog,
            text="Enter your Gmail address and 16-character App Password:",
            font=(self.FONT_FAMILY, 8),
            fg="#CBD5E1",
            bg=self.COLOR_PRIMARY
        ).pack(pady=(0, 14))

        form = tk.Frame(dialog, bg=self.COLOR_PRIMARY)
        form.pack(fill="x", padx=32)

        tk.Label(
            form,
            text="Sender Gmail:",
            font=(self.FONT_FAMILY, 9, "bold"),
            fg=self.TEXT_WHITE,
            bg=self.COLOR_PRIMARY
        ).grid(row=0, column=0, sticky="w", pady=6)

        email_entry = tk.Entry(
            form,
            font=(self.FONT_FAMILY, 10),
            bg=self.COLOR_CARD_BG,
            fg=self.TEXT_PRIMARY,
            bd=1,
            relief="solid",
            width=26
        )
        email_entry.grid(row=0, column=1, pady=6, ipady=4)

        tk.Label(
            form,
            text="App Password:",
            font=(self.FONT_FAMILY, 9, "bold"),
            fg=self.TEXT_WHITE,
            bg=self.COLOR_PRIMARY
        ).grid(row=1, column=0, sticky="w", pady=6)

        pass_entry = tk.Entry(
            form,
            font=(self.FONT_FAMILY, 10),
            show="•",
            bg=self.COLOR_CARD_BG,
            fg=self.TEXT_PRIMARY,
            bd=1,
            relief="solid",
            width=26
        )
        pass_entry.grid(row=1, column=1, pady=6, ipady=4)

        res = [None, None]

        def save_and_close():
            e = email_entry.get().strip()
            p = pass_entry.get().strip()
            if not e or not p:
                messagebox.showwarning("Input Required", "Please enter both Sender Email and App Password!", parent=dialog)
                return
            res[0] = e
            res[1] = p
            dialog.destroy()

        btn_box = tk.Frame(dialog, bg=self.COLOR_PRIMARY)
        btn_box.pack(fill="x", padx=32, pady=20)

        btn_send = tk.Button(
            btn_box,
            text=" Send Report ",
            font=(self.FONT_FAMILY, 9, "bold"),
            bg=self.COLOR_ACCENT,
            fg=self.TEXT_WHITE,
            activebackground=self.COLOR_ACCENT_HOVER,
            activeforeground=self.TEXT_WHITE,
            bd=0,
            cursor="hand2",
            command=save_and_close,
            padx=12,
            pady=7
        )
        btn_send.pack(side="left", expand=True, fill="x", padx=(0, 6))
        bind_hover(btn_send, self.COLOR_ACCENT, self.COLOR_ACCENT_HOVER)

        btn_cancel = tk.Button(
            btn_box,
            text=" Cancel ",
            font=(self.FONT_FAMILY, 9),
            bg="#334155",
            fg=self.TEXT_WHITE,
            activebackground="#475569",
            activeforeground=self.TEXT_WHITE,
            bd=0,
            cursor="hand2",
            command=dialog.destroy,
            padx=12,
            pady=7
        )
        btn_cancel.pack(side="right", expand=True, fill="x", padx=(6, 0))
        bind_hover(btn_cancel, "#334155", "#475569")

        dialog.bind("<Return>", lambda e: save_and_close())
        self.root.wait_window(dialog)
        return res[0], res[1]

    # ============================================================
    # 7. STUDENT REGISTRATION PAGE (PRESERVED)
    # ============================================================

    def build_register_page(self):
        split_frame = tk.Frame(self.page_register, bg=self.COLOR_MAIN_BG)
        split_frame.pack(fill="both", expand=True)

        # LEFT CARD: Registration Form & Interactive Step Workflow
        left_card = tk.LabelFrame(
            split_frame,
            text=" STUDENT REGISTRATION ",
            font=(self.FONT_FAMILY, 10, "bold"),
            fg=self.TEXT_PRIMARY,
            bg=self.COLOR_CARD_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=20,
            pady=16
        )
        left_card.pack(side="left", fill="both", expand=True, padx=(0, 16))

        tk.Label(
            left_card,
            text="Register a new student by capturing 30 face samples for LBPH recognition.",
            font=(self.FONT_FAMILY, 9),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG
        ).pack(anchor="w", pady=(0, 12))

        # Step Workflow Indicator Bar
        step_box = tk.Frame(
            left_card,
            bg=self.COLOR_MAIN_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=14,
            pady=8
        )
        step_box.pack(fill="x", pady=(0, 14))

        steps_text = "01 Enter Name  ➔  02 Start Camera  ➔  03 Capture Samples  ➔  04 Complete"
        tk.Label(
            step_box,
            text=steps_text,
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.COLOR_ACCENT,
            bg=self.COLOR_MAIN_BG
        ).pack(anchor="center")

        # Input: Student Full Name
        tk.Label(
            left_card,
            text="Student Full Name",
            font=(self.FONT_FAMILY, 9, "bold"),
            fg=self.TEXT_PRIMARY,
            bg=self.COLOR_CARD_BG
        ).pack(anchor="w", pady=(0, 5))

        self.reg_name_entry = tk.Entry(
            left_card,
            font=(self.FONT_FAMILY, 10),
            bg=self.COLOR_CARD_BG,
            fg=self.TEXT_PRIMARY,
            bd=1,
            relief="solid",
            highlightthickness=1,
            highlightbackground=self.COLOR_BORDER
        )
        self.reg_name_entry.pack(fill="x", pady=(0, 14), ipady=6)
        self.reg_name_entry.bind("<Return>", lambda e: self.start_registration())

        # Primary Button: Start Face Capture
        btn_capture = tk.Button(
            left_card,
            text="  📸  START FACE CAPTURE  ",
            font=(self.FONT_FAMILY, 10, "bold"),
            bg=self.COLOR_ACCENT,
            fg=self.TEXT_WHITE,
            activebackground=self.COLOR_ACCENT_HOVER,
            activeforeground=self.TEXT_WHITE,
            bd=0,
            cursor="hand2",
            pady=9,
            command=self.start_registration
        )
        btn_capture.pack(fill="x", pady=(0, 14))
        bind_hover(btn_capture, self.COLOR_ACCENT, self.COLOR_ACCENT_HOVER)

        # Live Progress Display Box during capture
        self.reg_progress_box = tk.Frame(
            left_card,
            bg=self.COLOR_MAIN_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=14,
            pady=10
        )
        self.reg_progress_box.pack(fill="x", pady=(0, 14))

        self.lbl_capture_status = tk.Label(
            self.reg_progress_box,
            text="READY TO CAPTURE",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_MAIN_BG
        )
        self.lbl_capture_status.pack(anchor="w")

        self.lbl_capture_counter = tk.Label(
            self.reg_progress_box,
            text="Progress: 0 / 30",
            font=(self.FONT_FAMILY, 8),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_MAIN_BG
        )
        self.lbl_capture_counter.pack(anchor="w", pady=(2, 6))

        self.progress_bar = ttk.Progressbar(
            self.reg_progress_box,
            style="Modern.Horizontal.TProgressbar",
            orient="horizontal",
            mode="determinate",
            maximum=30
        )
        self.progress_bar.pack(fill="x")

        # Registration Guidelines Card
        instr_card = tk.LabelFrame(
            left_card,
            text=" Registration Guidelines ",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.TEXT_SECONDARY,
            bg=self.COLOR_CARD_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=14,
            pady=10
        )
        instr_card.pack(fill="x")

        steps = [
            "• Enter student's full name in the text field above",
            "• Click START FACE CAPTURE to launch the camera",
            "• Keep face steady in front of the camera under good lighting",
            "• 30 face samples will be captured automatically",
            "• Press 'Q' inside the camera window to cancel capture"
        ]
        for step in steps:
            tk.Label(
                instr_card,
                text=step,
                font=(self.FONT_FAMILY, 8),
                fg=self.TEXT_PRIMARY,
                bg=self.COLOR_CARD_BG,
                justify="left"
            ).pack(anchor="w", pady=1)

        # RIGHT CARD: Registered Students Directory
        right_card = tk.LabelFrame(
            split_frame,
            text=" REGISTERED STUDENTS DIRECTORY ",
            font=(self.FONT_FAMILY, 10, "bold"),
            fg=self.TEXT_PRIMARY,
            bg=self.COLOR_CARD_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER,
            padx=20,
            pady=16
        )
        right_card.pack(side="right", fill="both", expand=True)

        reg_table_frame = tk.Frame(
            right_card,
            bg=self.COLOR_CARD_BG,
            bd=1,
            relief="solid",
            highlightbackground=self.COLOR_BORDER
        )
        reg_table_frame.pack(fill="both", expand=True)

        columns = ("name", "samples", "status")
        self.reg_tree = ttk.Treeview(
            reg_table_frame,
            columns=columns,
            show="headings",
            selectmode="browse",
            style="Custom.Treeview"
        )

        self.reg_tree.heading("name", text="Student Name")
        self.reg_tree.heading("samples", text="Face Samples")
        self.reg_tree.heading("status", text="Registration Status")

        self.reg_tree.column("name", width=220, anchor="w")
        self.reg_tree.column("samples", width=130, anchor="center")
        self.reg_tree.column("status", width=140, anchor="center")

        self.reg_tree.tag_configure("badge_registered", foreground=self.COLOR_ACCENT, font=(self.FONT_FAMILY, 9, "bold"))
        self.reg_tree.tag_configure("empty_msg", foreground=self.TEXT_MUTED, font=(self.FONT_FAMILY, 9, "italic"))

        reg_scrollbar = ttk.Scrollbar(reg_table_frame, orient="vertical", command=self.reg_tree.yview, style="Vertical.TScrollbar")
        self.reg_tree.configure(yscrollcommand=reg_scrollbar.set)

        self.reg_tree.pack(side="left", fill="both", expand=True)
        reg_scrollbar.pack(side="right", fill="y")

    def refresh_students_directory(self):
        for item in self.reg_tree.get_children():
            self.reg_tree.delete(item)

        if not os.path.exists(STUDENT_FOLDER):
            self.reg_tree.insert("", "end", values=("No registered students found", "-", "-"), tags=("empty_msg",))
            return

        dirs = sorted([d for d in os.listdir(STUDENT_FOLDER) if os.path.isdir(os.path.join(STUDENT_FOLDER, d))])
        for student_name in dirs:
            student_dir = os.path.join(STUDENT_FOLDER, student_name)
            sample_count = len([f for f in os.listdir(student_dir) if os.path.isfile(os.path.join(student_dir, f))])
            self.reg_tree.insert(
                "", "end",
                values=(student_name, f"{sample_count} Samples", "REGISTERED"),
                tags=("badge_registered",)
            )

        if len(dirs) == 0:
            self.reg_tree.insert("", "end", values=("No registered students yet", "-", "-"), tags=("empty_msg",))

        self.set_status(f"ℹ Directory Loaded — {len(dirs)} student(s) currently registered in database.")

    def start_registration(self):
        name = self.reg_name_entry.get().strip()
        if not name:
            messagebox.showwarning("Name Required", "Please enter Student Full Name first!")
            return

        student_path = os.path.join(STUDENT_FOLDER, name)
        if not os.path.exists(student_path):
            os.makedirs(student_path)

        detector = load_face_detector()
        camera = open_camera()

        if not camera.isOpened():
            messagebox.showerror("Camera Error", "Could not open camera device!")
            self.set_status("✕ Camera could not be opened.")
            return

        self.lbl_capture_status.configure(text="CAPTURING FACE SAMPLES", fg=self.COLOR_ACCENT)
        self.set_status(f"ℹ Capturing face samples for {name}... Look at camera.")
        sample_count = 0

        while True:
            success, frame = camera.read()
            if not success:
                break

            h_f, w_f = frame.shape[:2]
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(100, 100))

            # Modern HUD Header (Midnight Navy)
            cv2.rectangle(frame, (0, 0), (w_f, 48), (42, 23, 15), -1)
            cv2.putText(frame, f"SANJIVANI REGISTRATION - {name.upper()}", (15, 32),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)

            for (x, y, w, h) in faces:
                sample_count += 1
                face_img = gray[y:y + h, x:x + w]
                cv2.imwrite(os.path.join(student_path, f"{sample_count}.jpg"), face_img)

                # Update Tkinter Progressbar
                self.progress_bar['value'] = sample_count
                self.lbl_capture_counter.configure(text=f"Progress: {sample_count} / 30")
                self.root.update_idletasks()

                # Camera HUD Bounding Box & Text (Sky Cyan)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (248, 189, 56), 2)
                cv2.putText(frame, f"Samples: {sample_count} / 30", (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (248, 189, 56), 2)
                break

            # Bottom instruction ribbon
            cv2.rectangle(frame, (0, h_f - 32), (w_f, h_f), (42, 23, 15), -1)
            cv2.putText(frame, "Keep face centered | Press 'Q' to Cancel Capture", (15, h_f - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (184, 163, 148), 1)

            cv2.imshow("SANJIVANI ATTENDANCE — Student Registration", frame)

            if sample_count >= 30 or (cv2.waitKey(100) & 0xFF == ord('q')):
                break

        camera.release()
        cv2.destroyAllWindows()

        if sample_count >= 30:
            self.lbl_capture_status.configure(text="✓ STUDENT REGISTERED SUCCESSFULLY", fg=self.COLOR_SUCCESS)
            self.lbl_capture_counter.configure(text="Progress: 30 / 30 (Complete)")
            self.progress_bar['value'] = 30
            messagebox.showinfo("Registration Complete", f"✓ Student '{name}' registered successfully with 30 face samples!")
            self.reg_name_entry.delete(0, tk.END)
            self.refresh_students_directory()
            self.set_status(f"✓ Registered: {name} (30 face samples saved).")
        else:
            self.lbl_capture_status.configure(text="CAPTURE CANCELLED", fg=self.COLOR_ERROR)
            self.set_status(f"⚠ Registration cancelled for {name}.")

    # ============================================================
    # 8. NOTIFICATIONS & STATUS BAR
    # ============================================================

    def create_status_bar(self):
        status_frame = tk.Frame(self.root, bg=self.COLOR_PRIMARY, height=30)
        status_frame.pack(side="bottom", fill="x")
        status_frame.pack_propagate(False)

        self.status_var = tk.StringVar(value="ℹ System ready")
        status_bar = tk.Label(
            status_frame,
            textvariable=self.status_var,
            font=(self.FONT_FAMILY, 9),
            anchor="w",
            bg=self.COLOR_PRIMARY,
            fg="#CBD5E1",
            padx=18,
            pady=4
        )
        status_bar.pack(side="left", fill="x", expand=True)

        # Right version/hackathon tag
        tk.Label(
            status_frame,
            text="SANJIVANI AI ATTENDANCE v2.0  ",
            font=(self.FONT_FAMILY, 8, "bold"),
            fg=self.TEXT_MUTED,
            bg=self.COLOR_PRIMARY
        ).pack(side="right", padx=16)

    def set_status(self, msg):
        self.status_var.set(msg)
        self.root.update_idletasks()


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = SanjivaniAttendanceApp(root)
    root.mainloop()
