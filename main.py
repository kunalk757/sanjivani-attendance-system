# pyrefly: ignore [missing-import]
import cv2
import os
import sys
import time
import numpy as np
from datetime import datetime
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ============================================================
# CONFIGURATION
# ============================================================

STUDENT_FOLDER = "students"
ATTENDANCE_FILE = "attendance.csv"
ATTENDANCE_XLSX = "attendance.xlsx"
CASCADE_FILE = "haarcascade_frontalface_default.xml"
FACE_SIZE = (200, 200)

# LBPH confidence threshold for strict Unknown Face Rejection:
#   0 - 55  = Genuine Registered Student (Strong Match)
#   70+     = UNKNOWN FACE REJECTED (Access Denied)
CONFIDENCE_THRESHOLD = 70


# ============================================================
# STEP 1: Load the Haar Cascade for face detection
# ============================================================

def load_face_detector():
    """Load the Haar Cascade face detector."""
    cascade_path = CASCADE_FILE

    # Fallback to OpenCV's bundled cascade file if local copy is missing/empty
    if not os.path.exists(cascade_path) or os.path.getsize(cascade_path) == 0:
        cascade_path = cv2.data.haarcascades + CASCADE_FILE

    detector = cv2.CascadeClassifier(cascade_path)

    if detector.empty():
        print("ERROR: Face detector file not found.")
        print("Path:", cascade_path)
        exit()

    return detector


# ============================================================
# STEP 2: Train the LBPH recognizer from student face samples
# ============================================================

def train_recognizer(face_detector=None):
    """
    Scan students/ folder, load face images, apply histogram equalization,
    resize to FACE_SIZE, assign numeric labels, and train LBPHFaceRecognizer.

    Returns:
        recognizer: trained LBPHFaceRecognizer (or None if no data)
        label_to_name: dict mapping numeric label -> student name
    """
    faces = []
    labels = []
    label_to_name = {}
    name_to_label = {}
    current_label = 0

    if not os.path.exists(STUDENT_FOLDER):
        print(f"ERROR: '{STUDENT_FOLDER}' folder not found.")
        return None, {}

    # Process student subdirectories first (e.g. students/ROHIT/)
    dirs = [d for d in os.listdir(STUDENT_FOLDER) if os.path.isdir(os.path.join(STUDENT_FOLDER, d))]

    for student_name in sorted(dirs):
        student_dir = os.path.join(STUDENT_FOLDER, student_name)
        image_count = 0

        for img_name in sorted(os.listdir(student_dir)):
            img_path = os.path.join(student_dir, img_name)
            if not os.path.isfile(img_path):
                continue

            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if img is None or img.size == 0:
                continue

            # Resize to standard size & apply histogram equalization for lighting invariance
            img_resized = cv2.resize(img, FACE_SIZE)
            img_equalized = cv2.equalizeHist(img_resized)

            if student_name not in name_to_label:
                name_to_label[student_name] = current_label
                label_to_name[current_label] = student_name
                current_label += 1

            faces.append(img_equalized)
            labels.append(name_to_label[student_name])
            image_count += 1

        if image_count > 0:
            print(f"  Loaded {image_count} samples for student: {student_name}")

    # Process loose image files for students without a folder
    files = [f for f in os.listdir(STUDENT_FOLDER)
             if os.path.isfile(os.path.join(STUDENT_FOLDER, f)) and f.lower().endswith(('.jpg', '.jpeg', '.png'))]

    for file_name in sorted(files):
        student_name = os.path.splitext(file_name)[0]
        if student_name in name_to_label:
            continue

        file_path = os.path.join(STUDENT_FOLDER, file_name)
        img = cv2.imread(file_path, cv2.IMREAD_GRAYSCALE)
        if img is None or img.size == 0:
            continue

        # Extract face crop if full photo
        if face_detector is not None:
            detected_faces = face_detector.detectMultiScale(img, scaleFactor=1.1, minNeighbors=4, minSize=(50, 50))
            if len(detected_faces) > 0:
                (fx, fy, fw, fh) = detected_faces[0]
                img = img[fy:fy + fh, fx:fx + fw]

        if img.size > 0:
            img_resized = cv2.resize(img, FACE_SIZE)
            img_equalized = cv2.equalizeHist(img_resized)

            name_to_label[student_name] = current_label
            label_to_name[current_label] = student_name
            current_label += 1

            faces.append(img_equalized)
            labels.append(name_to_label[student_name])
            print(f"  Loaded 1 sample image for student: {student_name}")

    if len(faces) == 0:
        print("ERROR: No face samples found to train.")
        return None, {}

    # Create and train the LBPH recognizer
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.train(faces, np.array(labels))

    print(f"\nRecognizer trained with {len(faces)} images across {len(label_to_name)} student(s).\n")
    return recognizer, label_to_name


# ============================================================
# STEP 3: Attendance Storage Management (CSV + XLSX)
# ============================================================

def initialize_attendance_file():
    """Ensure both attendance.csv and attendance.xlsx exist with headers."""
    # CSV
    if not os.path.exists(ATTENDANCE_FILE) or os.path.getsize(ATTENDANCE_FILE) == 0:
        with open(ATTENDANCE_FILE, mode='w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Student Name", "Date", "Time", "Status"])

    # XLSX
    if not os.path.exists(ATTENDANCE_XLSX) or os.path.getsize(ATTENDANCE_XLSX) == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Student Name", "Date", "Time", "Status"])
        for col in range(1, 5):
            ws.cell(row=1, column=col).font = Font(bold=True)
        try:
            wb.save(ATTENDANCE_XLSX)
        except Exception:
            pass


def is_already_marked(student_name):
    """Check if student is already marked present for today in attendance.csv."""
    if not os.path.exists(ATTENDANCE_FILE):
        return False
    today_str = datetime.now().strftime("%Y-%m-%d")
    try:
        with open(ATTENDANCE_FILE, mode='r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if row and len(row) >= 1 and str(row[0]).strip().lower() == str(student_name).strip().lower():
                    if len(row) >= 4 and row[1] == today_str:
                        return True
                    elif len(row) == 2:
                        return True
    except Exception:
        pass
    return False


def mark_attendance(student_name):
    """Mark student attendance into both attendance.csv and attendance.xlsx."""
    initialize_attendance_file()
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%I:%M:%S %p")

    if is_already_marked(student_name):
        return False  # Already marked

    # Save to CSV
    try:
        with open(ATTENDANCE_FILE, mode='a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([student_name, today_str, time_str, "Present"])
    except Exception as e:
        print(f"WARNING: CSV save error: {e}")

    # Save to XLSX
    try:
        if os.path.exists(ATTENDANCE_XLSX) and os.path.getsize(ATTENDANCE_XLSX) > 0:
            wb = load_workbook(ATTENDANCE_XLSX)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Student Name", "Date", "Time", "Status"])

        ws.append([student_name, today_str, time_str, "Present"])
        wb.save(ATTENDANCE_XLSX)
    except Exception as e:
        print(f"WARNING: Excel save error: {e}")

    return True


# ============================================================
# STEP 4: Camera helper
# ============================================================

def open_camera():
    """Try opening camera using DirectShow backend first, then MSMF/ANY."""
    for idx in [0, 1, 2]:
        for backend in [cv2.CAP_DSHOW, cv2.CAP_MSMF, cv2.CAP_ANY]:
            cap = cv2.VideoCapture(idx, backend)
            if cap.isOpened():
                ret, frame = cap.read()
                if ret and frame is not None:
                    return cap
                cap.release()
    return cv2.VideoCapture(0)


# ============================================================
# STEP 5: Main attendance loop (Direct AI Face Recognition)
# ============================================================

def main():
    print("=" * 60)
    print("  SANJIVANI ATTENDANCE SYSTEM — AI FACE RECOGNITION")
    print("=" * 60)

    # Load face detector
    print("\nLoading face detector...")
    face_detector = load_face_detector()
    print("Face detector loaded.\n")

    # Train recognizer
    print("Training face recognizer...")
    recognizer, label_to_name = train_recognizer(face_detector)

    if recognizer is None:
        print("\nCannot start attendance — no trained faces found.")
        print("Run register.py first to register students.")
        return

    # Initialize storage files
    initialize_attendance_file()

    # Open camera
    camera = open_camera()
    if not camera.isOpened():
        print("ERROR: Camera could not be opened.")
        return

    print("Camera started. Look into the camera. Press Q to quit.\n")

    stop_program = False
    marked_student = None

    while True:
        success, frame = camera.read()
        if not success:
            print("Could not read from camera.")
            break

        h_f, w_f = frame.shape[:2]

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(100, 100))

        # Top Header Banner
        cv2.rectangle(frame, (0, 0), (w_f, 45), (15, 23, 42), -1)
        cv2.putText(frame, "SANJIVANI AI FACE ATTENDANCE", (15, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65, (74, 222, 128), 2)

        for (x, y, w, h) in faces:
            if w <= 0 or h <= 0 or y < 0 or x < 0:
                continue

            face_roi = gray[y:y + h, x:x + w]
            if face_roi is None or face_roi.size == 0:
                continue

            face_roi_resized = cv2.resize(face_roi, FACE_SIZE)
            face_roi_equalized = cv2.equalizeHist(face_roi_resized)

            label, confidence = recognizer.predict(face_roi_equalized)

            if confidence < CONFIDENCE_THRESHOLD:
                student_name = label_to_name.get(label, "Unknown")
                was_marked = mark_attendance(student_name)

                if was_marked:
                    display_text = f"MARKED PRESENT: {student_name}"
                    print(f"\n>>> ATTENDANCE MARKED: {student_name} (Confidence: {confidence:.1f})")
                    color = (0, 255, 0)
                else:
                    display_text = f"{student_name} (Already Marked Today)"
                    print(f"\n>>> INFO: {student_name} is already marked present today.")
                    color = (255, 200, 0)

                marked_student = student_name

                cv2.rectangle(frame, (x, y), (x + w, y + h), color, 3)
                cv2.putText(frame, display_text, (x, max(y - 15, 30)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.75, color, 2)

                cv2.imshow("Face Attendance System", frame)
                cv2.waitKey(2500)
                stop_program = True
                break
            else:
                student_name = "Unknown"
                color = (0, 0, 255)
                display_text = f"UNKNOWN FACE REJECTED ({confidence:.1f})"

                cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
                cv2.putText(frame, display_text, (x, max(y - 10, 25)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.65, color, 2)

        cv2.imshow("Face Attendance System", frame)

        if stop_program or (cv2.waitKey(1) & 0xFF == ord("q")):
            break

    camera.release()
    cv2.destroyAllWindows()

    print("\n" + "=" * 50)
    print("  SESSION SUMMARY")
    print("=" * 50)

    if marked_student:
        print(f"Attendance status processed for: {marked_student}")
    else:
        print("No attendance was marked this session.")

    print(f"Attendance records saved to:\n  - {ATTENDANCE_FILE}\n  - {ATTENDANCE_XLSX}")


if __name__ == "__main__":
    main()

